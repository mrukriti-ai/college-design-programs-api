import io
from typing import Dict, List
from sqlalchemy.orm import Session
from datetime import datetime
from openpyxl import load_workbook

from fastapi_app.database.models import College

REQUIRED_COLUMNS = [
    "name",
    "location_city",
    "location_country",
    "program_name",
    "program_type",
    "degree_level",
    "tuition_min",
    "tuition_max",
]

class ExcelImportService:
    @staticmethod
    def import_excel(file, db: Session) -> Dict:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        missing = [c for c in REQUIRED_COLUMNS if c not in headers]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

        col_index = {h: headers.index(h) for h in headers}

        inserted, updated, skipped = 0, 0, 0
        for row in ws.iter_rows(min_row=2):
            try:
                def get(col: str):
                    idx = col_index.get(col)
                    if idx is None:
                        return ""
                    val = row[idx].value
                    return val if val is not None else ""

                key_name = str(get("name")).strip()
                key_program = str(get("program_name")).strip()
                key_country = str(get("location_country")).strip()

                if not key_name or not key_program or not key_country:
                    skipped += 1
                    continue

                college = (
                    db.query(College)
                    .filter(
                        College.name == key_name,
                        College.program_name == key_program,
                        College.location_country == key_country,
                    )
                    .first()
                )

                data = {
                    "name": key_name,
                    "location_city": str(get("location_city")).strip(),
                    "location_country": key_country,
                    "program_name": key_program,
                    "program_type": str(get("program_type")).strip(),
                    "degree_level": str(get("degree_level")).strip(),
                    "tuition_min": ExcelImportService._to_float(get("tuition_min")),
                    "tuition_max": ExcelImportService._to_float(get("tuition_max")),
                    "application_deadline": ExcelImportService._parse_date(get("application_deadline")),
                    "program_description": str(get("program_description")).strip(),
                    "admission_requirements": str(get("admission_requirements")).strip(),
                    "contact_email": str(get("contact_email")).strip(),
                    "website_url": str(get("website_url")).strip(),
                }

                if college:
                    for k, v in data.items():
                        setattr(college, k, v)
                    updated += 1
                else:
                    db.add(College(**data))
                    inserted += 1

            except Exception:
                skipped += 1
                continue

        db.commit()
        return {"inserted": inserted, "updated": updated, "skipped": skipped}

    @staticmethod
    def _to_float(value):
        try:
            if value in (None, "", "nan"):
                return None
            return float(value)
        except Exception:
            return None

    @staticmethod
    def _parse_date(value):
        if value in (None, "", "nan"):
            return None
        if isinstance(value, datetime):
            return value.date()
        try:
            return datetime.strptime(str(value), "%Y-%m-%d").date()
        except Exception:
            try:
                return datetime.strptime(str(value), "%d/%m/%Y").date()
            except Exception:
                return None 